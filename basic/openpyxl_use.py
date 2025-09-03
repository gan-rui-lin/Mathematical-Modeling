from openpyxl import load_workbook

file_path = "example.xlsx"

wb = load_workbook(file_path)

ws = wb["问题2的订购方案结果"]

cell_range = "B7:K26"

cell_block = ws[cell_range]

print(cell_block)

num_rows = len(cell_block)
num_cols = len(cell_block[0]) if num_rows > 0 else 0

print(num_rows, num_cols)

import numpy as np

fill_value = 888
fill_data = np.full((num_rows, num_cols), fill_value, dtype=np.float64)

fill_data[1, 1] = np.nan

for i, row in enumerate(cell_block):
    # print(f"the {i}th row is {row}")
    for j, cell in enumerate(row):
        # print(f"the {j}th cell is {cell}")
        if fill_data[i, j] is not None:
            cell.value = fill_data[i, j]

wb.save("filled_example.xlsx")

